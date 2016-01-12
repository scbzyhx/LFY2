#
# -*- coding:utf-8

from openpyxl import load_workbook
import xlrd

NUMBER = 1
STRING = 2
NONE = 3

class YXlrdWorkbook(object):
    EMPTY = 0
    STRING = 1
    NUMBER = 2
    DATE = 3
    BOOLEAN = 4
    ERROR = 5
    def __init__(self,wb,**kwargs):
        super(YXlrdWorkbook,self).__init__()
        self.wb = wb
        self.sheet = None
        
    def get_sheet_names(self):
        return self.wb._sheet_names
    def pin_sheet_by_name(self,name):
        self.sheet = self.wb.sheet_by_name(name)
    #simple version
    def _covert_pos(self,pos):
        col = pos[0]
        #if is_number(pos[1:]):
        row = pos[1:]
        return  int(row), ord(col) - ord('A')
    def get_cell_type(self,pos):
        r,c = self._covert_pos(pos)
        return self.sheet.cell(r,c).ctype
         
        
    def get_cell_value(self,pos):
        r,c = self._covert_pos(pos)
        #print r,"  " ,c
        return self.sheet.cell(r,c).value
    def get_nrows(self):
        return self.sheet.nrows
    def get_ncols(self):
        return self.sheet.ncols
        
    def test(self):
        ce = self.sheet.cell(0,0)
        print dir(self.sheet)
        print dir(ce)
        
class YOpenpyxlWorkbook(object):
    EMPTY = 0
    STRING = 1
    NUMBER = 2
    DATE = 3
    BOOLEAN = 4
    ERROR = 5
    def __init__(self,wb,**kwargs):
        super(YOpenpyxlWorkbook,self).__init__()
        self.sheet = None
        self.wb = wb
        
    def get_sheet_names(self):
        return self.wb.get_sheet_names()
    def pin_sheet_by_name(self,name):
        self.sheet = self.wb.get_sheet_by_name(name)
        
    def get_cell_type(self,pos):
        return 
    def get_cell_value(self,pos):
        return self.sheet[pos]
    def get_nrows(self):
        raise NotImplementedError
    def get_ncols(self):
        raise NotImplementedError
    #wb = load_workbook(path, use_iterators=True)
    #    sheet = wb.worksheets[0]

    #    row_count = sheet.max_row
    #    column_count = sheet.max_column

def YRWorkbook(filename):
     if filename.split(".")[-1] == "xls":
         return YXlrdWorkbook(xlrd.open_workbook(filename, encoding_override="gbk"))
            
     elif filename.split(".")[-1] == "xlsx":
         return YOpenpyxlWorkbook(load_workbook(filename))
     else:
         assert 0, "%s is an wrong file" % filename 

    

if __name__=="__main__":
    tb = YRWorkbook("C:\\Users\\yanghaixiang\\Documents\\GitHub\\TPI_VC70160106A07.xls")
    tb.pin_sheet_by_name("TPI1")
    print tb.get_cell_value("A13")
    print tb.get_cell_value("B13")
    print tb.get_cell_value("E13")
    print tb.get_cell_value("F13")
    print tb.get_cell_value("G13")
    print tb.get_cell_type("K13")
    print tb.get_cell_value("K13")
    
    
