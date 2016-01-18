#
# -*- coding:utf-8 -*-
import logging

from openpyxl import load_workbook, Workbook
import utils


class HAWB:
    CONTRACT_ID_COL = "A"
    HAWB_ID_COL = "B"
    AMOUNT_COL = "C"
    H_CONTRACT_ID = u"合同号"
    H_HAWB_ID = u"HWAB"
    H_AMOUNT = u"总值"
    def __init__(self,filename,**kwargs):
        self.filename = filename
        #self.row = 0
        try:
            self.wb = load_workbook(filename)
            self.sheet = self.wb.active
            #self.row = self.sheet.get_highest_row()
        except IOError:
            self.wb = Workbook()
            self.sheet = self.wb.active
            #self.row
        self.row = self.sheet.get_highest_row()
        #print self.row
        self.row += 1
        if self.row == 1:
            self.insert_head()
    def insert_record(self,c_id,hawb, amount):
        self.sheet[self.CONTRACT_ID_COL+str(self.row)] = c_id
        self.sheet[self.HAWB_ID_COL + str(self.row)] = hawb
        #self.sheet[self.IN_PRICE_COL + str(self.row)] = in_price
        #self.sheet[self.OUT_PRICE_COL + str(self.row)] = out_price
        self.sheet[self.AMOUNT_COL + str(self.row)] = amount#contract_price
        self.row += 1
    def insert_head(self):
        self.insert_record(self.H_CONTRACT_ID, self.H_HAWB_ID, self.H_AMOUNT)
    def save(self):
        self.wb.save(self.filename)
class Account:
    CONTRACT_ID_COL = "A"
    PN_COL = "B"
    IN_PRICE_COL = "C"
    OUT_PRICE_COL = "D"
    CONTRACT_PRICE_COL = "E"
    
    H_CONTRACT_ID = u"合同号"
    H_PN = u"P/N"
    H_IN_PRICE = u"进口价格"
    H_OUT_PRICE =u"出口价格"
    H_CONTRACT_PRICE = u"合同价格"
    
    def __init__(self,filename,**kwargs):
        self.filename = filename
        #self.row = 0
        try:
            self.wb = load_workbook(filename)
            self.sheet = self.wb.active
            #self.row = self.sheet.get_highest_row()
        except IOError:
            self.wb = Workbook()
            self.sheet = self.wb.active
            #self.row
        self.row = self.sheet.get_highest_row()
        #print self.row
        self.row += 1
        if self.row == 1:
            
            self.insert_head()
        
    def insert_record(self,c_id,pn, in_price, out_price, contract_price):
        self.sheet[self.CONTRACT_ID_COL+str(self.row)] = c_id
        self.sheet[self.PN_COL + str(self.row)] = pn
        self.sheet[self.IN_PRICE_COL + str(self.row)] = in_price
        self.sheet[self.OUT_PRICE_COL + str(self.row)] = out_price
        self.sheet[self.CONTRACT_PRICE_COL + str(self.row)] = contract_price
        self.row += 1
        
    def insert_head(self):
        self.sheet[self.CONTRACT_ID_COL+str(self.row)] = self.H_CONTRACT_ID
        self.sheet[self.PN_COL + str(self.row)] = self.H_PN
        self.sheet[self.IN_PRICE_COL + str(self.row)] = self.H_IN_PRICE
        self.sheet[self.OUT_PRICE_COL + str(self.row)] = self.H_OUT_PRICE
        self.sheet[self.CONTRACT_PRICE_COL + str(self.row)] = self.H_CONTRACT_PRICE
        self.row += 1 #start from 2 now 
    def save(self):
        self.wb.save(self.filename)

if __name__=="__main__":
    hawb = Account(".\\workspace\\abc.xlsx")
    hawb.insert_record("abcd","aaa",1,2,3)
    hawb.save()
