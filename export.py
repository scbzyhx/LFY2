#
# -*- coding:utf-8
import logging
import os
import utils
from openpyxl import load_workbook

class Export:
    HEADER = "P/N"
    def __init__(self,filename, **kwargs):
        self.filename = filename
        try:
            self.wb = load_workbook(filename)
        except IOError:
            utils.ABT(u"ERROR: 出口文件缺失")
        self.db = {}#[0,0],,,max_price, qty
        self._process()
        
    def _process(self):
        names = self.wb.get_sheet_names()
        
        for name in names:
            #print name
            ws = self.wb.get_sheet_by_name(name)
            for row in xrange(1,ws.max_row +1):
                
                pn = ws["A"+str(row)].value
                qty = ws["B"+str(row)].value
                uprice = ws["C"+str(row)].value
                if pn is None or qty is None or uprice is None:
                    continue
                pn.strip()
                
                
                if pn == self.HEADER:
                    continue
                    
                qty = int(qty)
                uprice = float(uprice)
                if qty <= 0 or uprice <=0:
                    utils.ABT(u"ERROR: 出口表中，这条记录(pn=%s,qty=%d,unit_price=%f) 有错误"%(pn,qty,uprice))
                
                if not self.db.has_key(pn):
                    self.db[pn] = [0,0] 
                self.db[pn][0] = max(self.db[pn][0],uprice)
                self.db[pn][1] += qty
                
                #print pn,qty,uprice
                #print type(pn)," ", type(qty), "  ",type(uprice)
    def get_unit_price(self,pn):
        if self.db.has_key(pn):
            return self.db[pn][0]
    def get_qty(self,pn):
        if self.db.has_key(pn):
            return self.db[pn][1]
    def get_pns(self):
        return self.db.keys()
            
if __name__=="__main__":
    ex = Export(u"C:\\Users\\yanghaixiang\\Desktop\\一键制单\\出口单价与数量.xlsx")
    print ex.db 
    print ex.get_unit_price("MJVE2CH/A")
    print ex.get_pns()
    
