#
# -*- coding:utf-8
from openpyxl import load_workbook
import logging
import utils

class YInfoDB(object):
    HEADER = u"料号"
    MODEL = "model"
    CODE = "code"
    DESC = "desc"
    def __init__(self,filename,**kwargs):
        self.filename = filename
        self.logger = logging.getLogger("InfoDB")
        try:
            self.wb = load_workbook(filename,use_iterators=True)
        except IOError:
            utils.ABT(u"规范申报数据库文件缺失")
        try:
            self.sheet = self.wb.get_sheet_by_name("DATA")
        except KeyError:
            utils.ABT(u"规范申报数据库中缺失 DATA 表单")
            
        self.data = {} #each is a dict
        for row in xrange(1,self.sheet.max_row+1):
            pn = self.sheet["A"+str(row)].value
            model = self.sheet["B"+str(row)].value
            code = self.sheet["C"+str(row)].value
            desc = self.sheet["D"+str(row)].value
            if pn is None or model is None or code is None or desc is None:
                self.logger.debug(u"在规范数据库缺失")
                continue
            pn.strip()
            model.strip()
            desc.strip()
            if pn == self.HEADER:
                continue
            if self.data.has_key(pn):
                #if self.data[pn][self.MODEL] == model and self.data[pn][self.CODE] == code and self.data[pn][self.DESC] == desc:
                continue
                """
                else:
                    print("duplicate PN(%s) in database file" % pn)
                    print "model=", model
                    print self.data[pn][self.MODEL]
                    print "code=",code
                    print self.data[pn][self.CODE]
                    print "desc=", desc
                    print self.data[pn][self.DESC]
                    assert 0
                """
            #assert not self.data.has_key(pn), (u"(%s) is duplicate in database table(%s)" %(pn,self.filename))
            
            self.data[pn] = {}
            self.data[pn][self.MODEL] = model
            self.data[pn][self.CODE] = code
            self.data[pn][self.DESC] = desc
            
            
        
        
        
    def get_model(self,pn):
        if self.data.has_key(pn):
            return self.data[pn][self.MODEL]
        
    def get_code(self,pn):
        if self.data.has_key(pn):
            return self.data[pn][self.CODE]
        
    def get_desc(self,pn):
        if self.data.has_key(pn):
            return self.data[pn][self.DESC]
        

if __name__=="__main__":
    infodb = YInfoDB(u"C:\\Users\\yanghaixiang\\Desktop\\一键制单\\规范申报数据库.xlsx")
    print infodb.get_model(u"Z0NP001WF")
    print infodb.get_code(u"Z0NP001WF")
    print infodb.get_desc(u"Z0NP001WF")
    
    
