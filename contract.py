#
# -*- coding:utf-8 -*-
from openpyxl.styles import Border,Side,Alignment, Protection, Font,colors
import utils

class Contract(object):
    QTY_COL = "A"
    DESC_COL = "B"
    PRICE_COL = "F"
    AMOUNT_COL = "G"
    ROW_START = 17
    ROW_END = 37

    CONTRACT_ID_POS = "G4"
    CONTRACT_DATE_POS = "G8"
    
    TOTAL_QTY_POS = "A37"
    TOTAL_AMOUNT_POS = "G37"
    
    def __init__(self,sheet,**kwargs):
    	super(Contract,self).__init__()
        self.sheet = sheet #complete here
        self.row = 17

    def insert_record(self,qty,desc,price,amount):
        if self.row >= self.ROW_END:
    		return -1
    	desc_cell = self.sheet[self.DESC_COL+str(self.row)]
        desc_cell_border = desc_cell.border
        #print "border= \n\n",desc_cell_border
        
        self.sheet[self.QTY_COL+str(self.row)] = qty
        self.sheet[self.DESC_COL+str(self.row)] = desc 
        self.sheet[self.PRICE_COL+str(self.row)] = price
        self.sheet[self.AMOUNT_COL+str(self.row)] = amount

        #OK, future
        
        desc_cell.alignment = Alignment(horizontal='left')
        desc_cell.border = desc_cell_border
    	self.row += 1
        return 0

    def insert_contract_id(self,c_id):
    	self.sheet[self.CONTRACT_ID_POS] = c_id
        

    def insert_contract_date(self,date):
        self.sheet[self.CONTRACT_DATE_POS] = date

    def close(self):
        "save"
        pass
        
    def get_cell(self,pos=QTY_COL+str(ROW_START)):
    	return self.sheet[pos].value#.encode('ascii','ignore')
    def merged_cells(self):
        self.sheet.merge_cells('B17:D17')
    def fix_border(self):
        pass
    def set_total_amount(self,total):
        self.sheet[self.TOTAL_AMOUNT_POS] = total
    def set_total_qty(self,total_qty):
        self.sheet[self.TOTAL_QTY_POS] = total_qty
    def fix_borders(self):
        head_border_mid = Border(top=Side(border_style="thin",color=colors.BLACK),bottom=Side(border_style="medium",color=colors.BLACK))
        head_border_right = Border(top=Side(border_style="thin",color=colors.BLACK),bottom=Side(border_style="medium",color=colors.BLACK),right=Side(border_style="thin",color=colors.BLACK))
        
        border_mid = Border(top=Side(border_style="thin",color=colors.BLACK),bottom=Side(border_style="thin",color=colors.BLACK))
        border_right = Border(top=Side(border_style="thin",color=colors.BLACK),bottom=Side(border_style="thin",color=colors.BLACK),right=Side(border_style="thin",color=colors.BLACK))
        self.sheet["C16"].border = head_border_mid
        self.sheet["D16"].border = head_border_right;
        for i in xrange(self.ROW_START,self.ROW_END):
            self.sheet["C"+str(i)].border = border_mid            
            self.sheet["D"+str(i)].border = border_right
        
        
    def remove_empty_rows(self): #copy
        #test 
        #self.row = 22
        last_row = self.sheet.get_highest_row() 
        #last_col = 8#self.sheet.get_highest_column() + 1
        #print last_col
        for i in xrange(self.row,self.ROW_END):
            self.sheet.row_dimensions[i].hidden = True
            #for j in xrange(last_col):
            #    old_pos = chr(j+ord("A"))+str(i)
            #    new_pos = chr(j+ord("A")) +str(self.row)
            #    print old_pos, new_pos
            #    self.sheet[new_pos] = self.sheet[old_pos].value
            #    self.sheet[new_pos].font = self.sheet[old_pos].font
            #    self.sheet[new_pos].fill = self.sheet[old_pos].fill
            #    self.sheet[new_pos].border = self.sheet[old_pos].border
            #    self.sheet[new_pos].alignment = self.sheet[old_pos].alignment
                #print self.sheet[old_pos].value
            #self.row += 1
                
    def print_border(self):
        """
        left = border.copy(right=None)
        right = border.copy(left=Side(border_style=None))
        middle = border.copy(left=Side(border_style=None), right=Side(border_style=None))
        st = ws['b2'].style
        ws['b2'].style = st.copy(border=left)
        ws['d2'].style = st.copy(border=middle)
        ws['e2'].style = st.copy(border=middle)
        ws['f2'].style = st.copy(border=right)
        """
        desc_cell = self.sheet["B16"]
        desc_cell_border = desc_cell.border
        print self.sheet["B16"].border
        print desc_cell_border



if __name__ == "__main__":
    print u"人是人，鬼是鬼"
    from openpyxl import load_workbook
    try:
    	wb = load_workbook(filename=u"C:\\Users\\yanghaixiang\\Desktop\\一键制单\\sample.xlsx")
    	contract = Contract(wb.get_sheet_by_name(u"合同"))
    	#contract.show_cell("A16")
        print contract.sheet.get_highest_row()
        print dir(contract.print_border())
        contract.remove_empty_rows()
    	#contract.print_border(u"C:\\Users\\yanghaixiang\\Desktop\\一键制单\\sample_move.xlsx")
        wb.save(u"C:\\Users\\yanghaixiang\\Desktop\\一键制单\\sample_move.xlsx")
    	#for row in contract.sheet.iter_rows("A1:A4"):
        #    print type(row)
        #    print dir(row)
    except IOError :
    	print "err"
